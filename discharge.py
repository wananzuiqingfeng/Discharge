"""
    充放电数据快速自动化录入 Excel 报告文件
"""


__Auther__ = 'Super Pig'
__Version__ = '1.3'


import os
import time
import shutil
import pprint

from datetime import datetime

import openpyxl


# BatteryMon 工具生成的数据文件可能包含的多余信息
USELESS_TEXT = 'Unique ID, Manufacturer, Chemistry, Voltage (Volts)'

# 收集几个循环的数据，默认为 3
CYCLES_NUMBER = 3

# 数据项计数，正确数量为 11
DATA_ITEM_NUMBER = 11

# 充放电测试报告文件名
EXCEL_FILE_NAME = './充放电测试报告.xlsx'

# 所有数据文件的存储位置
ALL_DATA_FILE_STORAGE_PATH = "./Result/"


# 装饰器: 输出程序运行时间
def timing(func):
    def wrapper(*args, **kw):
        start_time = time.time()
        print(f"Start Time: {time.asctime(time.localtime(start_time))}\n")

        func(*args, **kw)

        end_time = time.time()
        print()
        print(f"End Time: {time.asctime(time.localtime(end_time))}")
        print(f"Total Time: {end_time - start_time:.6f}s\n")

    return wrapper


class Discharge(object):
    """充放电数据快速统计收集并录入Excel文件

    Attributes:
        charge_files: 充电数据文件名列表
        discharge_files: 放电数据文件名列表

        charge_datas: 存放清洗过的充电数据的容器
        discharge_datas: 存放清洗过的放电数据的容器

        useless_text: 额外的日志信息(理应舍弃)
        cycles：收集几个循环的计数，默认为 3，特殊情况下该数字可能为 2，支持被修改，目前不支持 cycles 大于 3
        data_item_number: 数据项计数，应为 11
        excel_file: 存储数据的 Excel 文件名
        move_files: 需要移动的文件名列表，程序结束后会将充放电数据文件移动(避免目录臃肿以及支持下一台设备的数据录入)

    Months:
        collect_charge_data: 收集充电数据
        collect_discharge_data: 收集放电数据
        save_data: 将数据存储至容器
        save_to_excel: 将数据存储至 Excel 表格

        inspection_data_file: 检查充放电数据文件是否正确存放
        open_signal_switch: 打开信号开关
        calculate_time: 计算充放电过程总耗时

        main: 调度程序运行
    """

    def __init__(self):
        # 充放电数据文件名列表
        self.charge_files = []
        self.discharge_files = []

        # 充放电数据列表
        self.charge_datas = []
        self.discharge_datas = []

        # 多余信息
        self.useless_text = USELESS_TEXT

        # 几个循环的数据
        self.cycles = CYCLES_NUMBER

        # 数据项计数
        self.data_item_number = DATA_ITEM_NUMBER

        # 存储数据的Excel文件名
        self.excel_file = EXCEL_FILE_NAME

        # 需要移动的文件
        self.move_files = []

    def inspection_data_file(self):
        """检查充放电数据文件是否正确存放"""
        print('正在检测充放电数据文件是否存放正确: ')

        if self.cycles > 3:
            raise ValueError("目前暂不支持大于 3 个循环")

        for file in os.listdir('./'):
            file = file.lower()		# 用户可能使用大写
            if file.endswith('txt'):
                if file.startswith('c'):
                    self.charge_files.append(file)
                elif file.startswith('f'):
                    self.discharge_files.append(file)

        # 检查数量是否正确
        charge_files_number = len(self.charge_files)
        dischatge_files_number = len(self.discharge_files)

        if charge_files_number < self.cycles:
            raise ValueError('充电数据文件数量异常.\n')

        if dischatge_files_number < self.cycles:
            raise ValueError('放电数据文件数量异常.\n')

        # 扩展 move_files 列表，以便移动文件时使用
        self.move_files.extend(self.charge_files)
        self.move_files.extend(self.discharge_files)

        # 将多余循环的数据舍弃(如果有)
        for _ in range(charge_files_number - self.cycles):
            self.charge_files.pop(0)

        for _ in range(dischatge_files_number - self.cycles):
            self.discharge_files.pop(0)

        print('True!\n')

    def open_signal_switch(self):
        """打开信号开关"""
        # 控制是否收集在低电至满电之间第一次达到指定电量(低于 10%，10% ... 100%)时的充电数据 (反之亦然)
        self.low_power = True
        self.electricity_10  = True
        self.electricity_20  = True
        self.electricity_30  = True
        self.electricity_40  = True
        self.electricity_50  = True
        self.electricity_60  = True
        self.electricity_70  = True
        self.electricity_80  = True
        self.electricity_90  = True
        self.electricity_100 = True

    def save_data(self, datas, str_text_list):
        """将数据存储至 datas 容器

        Args:
            datas: 存储 dict 数据的列表 -> list
            str_text_list: 字符串文本组成的列表 -> list
                E.g:
                ['2022-04-07', ' 21:33:18', ' OK', ' 5%', ' 0.00', ' -1', ' 11100', ' ', ' 7.596',
                ' 11100', ' 5.0%', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
        """
        # 清洗、收集目标数据
        datas.append({
            'date': str_text_list[0],
            'time': str_text_list[1].strip(),
            'electric_quantity': str_text_list[3]
        })

    def collect_charge_data(self):
        """收集充电数据"""
        for file in self.charge_files:
            # 初始化临时存放数据的容器以及打开控制信号开关
            datas = []
            self.open_signal_switch()

            with open(file, 'r', encoding='utf-8') as f:
                print(f'正在从 {file} 文件中收集充电数据: ')
                str_text_lists = f.readlines()

            # 清洗 BatteryMon 工具多余的信息(如果有)
            if self.useless_text in str_text_lists[0]:
                str_text_lists = str_text_lists[4:]

            for str_text in str_text_lists:
                str_text_list = str_text.split(',')

                # 根据信号开关指示收集数据
                if self.low_power and int(str_text_list[3][1:2]) < 10:		# 默认小于 10 为低电
                    self.save_data(datas, str_text_list)
                    self.low_power = False

                elif self.electricity_10 and str_text_list[3] == ' 10%':
                    self.save_data(datas, str_text_list)
                    self.electricity_10 = False

                elif self.electricity_20 and str_text_list[3] == ' 20%':
                    self.save_data(datas, str_text_list)
                    self.electricity_20 = False

                elif self.electricity_30 and str_text_list[3] == ' 30%':
                    self.save_data(datas, str_text_list)
                    self.electricity_30 = False

                elif self.electricity_40 and str_text_list[3] == ' 40%':
                    self.save_data(datas, str_text_list)
                    self.electricity_40 = False

                elif self.electricity_50 and str_text_list[3] == ' 50%':
                    self.save_data(datas, str_text_list)
                    self.electricity_50 = False

                elif self.electricity_60 and str_text_list[3] == ' 60%':
                    self.save_data(datas, str_text_list)
                    self.electricity_60 = False

                elif self.electricity_70 and str_text_list[3] == ' 70%':
                    self.save_data(datas, str_text_list)
                    self.electricity_70 = False

                elif self.electricity_80 and str_text_list[3] == ' 80%':
                    self.save_data(datas, str_text_list)
                    self.electricity_80 = False

                elif self.electricity_90 and str_text_list[3] == ' 90%':
                    self.save_data(datas, str_text_list)
                    self.electricity_90 = False

                elif self.electricity_100 and str_text_list[3] == ' 100%':
                    self.save_data(datas, str_text_list)
                    self.electricity_100 = False

            if len(datas) == self.data_item_number:
                self.charge_datas.extend(datas)
                print('充电数据收集成功!\n')
                pprint.pprint(datas)
                print()
            else:
                print('充电数据收集异常!\n')
                pprint.pprint(datas)
                print()

    def collect_discharge_data(self):
        """收集放电数据"""
        for file in self.discharge_files:
            datas = []
            self.open_signal_switch()

            with open(file, 'r', encoding='utf-8') as f:
                print(f'正在从 {file} 文件中收集放电数据: ')
                str_text_lists = f.readlines()

            if self.useless_text in str_text_lists[0]:
                str_text_lists = str_text_lists[4:]

            # 如果放电数据文件的末尾没有低于 6% 电量的数据，则添加一行这样的数据保证程序稳定运行
            if int(str_text_lists[-2].split(",")[3][1:2]) == 6:
                new_text = str_text_lists[-1].replace("6%", "5%")
                str_text_lists.append(new_text)

            for str_text in str_text_lists:
                str_text_list = str_text.split(',')

                if self.electricity_100 and str_text_list[3] == ' 100%':
                    self.save_data(datas, str_text_list)
                    self.electricity_100 = False

                elif self.electricity_90 and str_text_list[3] == ' 90%':
                    self.save_data(datas, str_text_list)
                    self.electricity_90 = False

                elif self.electricity_80 and str_text_list[3] == ' 80%':
                    self.save_data(datas, str_text_list)
                    self.electricity_80 = False

                elif self.electricity_70 and str_text_list[3] == ' 70%':
                    self.save_data(datas, str_text_list)
                    self.electricity_70 = False

                elif self.electricity_60 and str_text_list[3] == ' 60%':
                    self.save_data(datas, str_text_list)
                    self.electricity_60 = False

                elif self.electricity_50 and str_text_list[3] == ' 50%':
                    self.save_data(datas, str_text_list)
                    self.electricity_50 = False

                elif self.electricity_40 and str_text_list[3] == ' 40%':
                    self.save_data(datas, str_text_list)
                    self.electricity_40 = False

                elif self.electricity_30 and str_text_list[3] == ' 30%':
                    self.save_data(datas, str_text_list)
                    self.electricity_30 = False

                elif self.electricity_20 and str_text_list[3] == ' 20%':
                    self.save_data(datas, str_text_list)
                    self.electricity_20 = False

                elif self.electricity_10 and str_text_list[3] == ' 10%':
                    self.save_data(datas, str_text_list)
                    self.electricity_10 = False

                # 末尾可能需要做特殊处理
                # 设备在完成放电关机后再次开机时测试程序依旧在运行，会记录开机后的电源信息
                # 筛选出开机时所记录数据上一行的数据才是最后放电时间的数据
                # tips: 开机记录的数据并不一定是 5% 电量(低于 5% 即未做低电保护)
                elif self.low_power and int(str_text_list[3][:-1]) < 6:
                    index = str_text_lists.index(str_text) - 1
                    str_text_list = str_text_lists[index].split(',')
                    self.save_data(datas, str_text_list)
                    self.low_power = False

            if len(datas) == self.data_item_number:
                self.discharge_datas.extend(datas)
                print('放电数据收集成功!\n')
                pprint.pprint(datas)
                print()
            else:
                print('放电数据收集异常!\n')
                pprint.pprint(datas)
                print()

    def save_to_excel(self):
        """将数据存储至 Excel 报告文件"""
        # 根据当前时间生成本次操作的文件夹名称
        path = ALL_DATA_FILE_STORAGE_PATH + time.strftime("%Y%m%d-", time.localtime())

        # 保证同一时间(精确到分)多次操作生成的名称正确计数
        count = 1
        path += f"({str(count)})/"
        while True:
            if os.path.exists(path):
                if len(os.listdir(path)) == 3:
                    path = path.replace(f"({str(count)})", f"({str(count + 1)})")
                    count += 1
                else:
                    break
            else:
                os.makedirs(path)
                shutil.copy(self.excel_file, path)
                break

        self.excel_file = path + self.excel_file[2:]
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb.active

        # 根据设备 #1 的数据是否空白来判断当前数据是那台设备的
        device_id = 1 if sheet.cell(4, 7).value is None else 2

        print(f'正在将设备 #{device_id} 的数据录入 {self.excel_file} 文件.\n')

        # 根据设备编号选择对应的行
        discharge_row = 4 if device_id == 1 else 7
        charge_row = 13 if device_id == 1 else 16

        # 默认收集 3 个循环的数据，由 self.cycles 控制
        for _ in range(self.cycles):
            column = 7

            for number in range(self.data_item_number):
                # dict E.g:
                # {'date': '2022-04-07', 'time': '21:33:18', 'electric_quantity': ' 5%'}
                charge_dict = self.charge_datas.pop(0)
                charge_time = charge_dict['time'][:-3]

                discharge_dict = self.discharge_datas.pop(0)
                discharge_time = discharge_dict['time'][:-3]

                # save
                sheet.cell(charge_row, column).value = charge_time
                sheet.cell(discharge_row, column).value = discharge_time

                # 下一列
                column += 1

                # 记录开始的日期时间
                if number == 0:
                    start_charge_date = charge_dict['date']
                    start_charge_time = charge_time

                    start_discharge_date = discharge_dict['date']
                    start_discharge_time = discharge_time

            # 写入充电时长
            # 充电耗时
            result = self.calculate_time(start_charge_date, start_charge_time,
                                         charge_dict['date'], charge_time)    # 循环的最后是结束的日期时间
            sheet.cell(charge_row, column).value = result

            # 放电耗时
            result = self.calculate_time(start_discharge_date, start_discharge_time,
                                         discharge_dict['date'], discharge_time)
            sheet.cell(discharge_row, column).value = result

            # 下一行
            charge_row += 1
            discharge_row += 1

        try:
            wb.save(self.excel_file)
        except:
            pass
        else:
            print(f'充放电数据成功存储至 {self.excel_file} 文件.\n')

            # 移动充放电数据文件到指定目录下
            # 报告默认为两台设备的数据，因此需要将两次数据分开存放，根据 device_id 计数即可
            path += str(device_id)
            os.makedirs(path)
            for file in self.move_files:
                shutil.move(file, path)

    def calculate_time(self, start_date, start_time, end_date, end_time):
        """计算总耗时

        Args:
            start_date: 起始日期
            start_time: 起始时间
            end_date: 结束日期
            end_time: 结束时间

        Return:
            text: 一段固定格式的文本，例如: 4h14min
        """
        # 开始/结束的日期时间列表，最终格式例如: [2001, 11, 25, 12, 30] -> 年、月、日、时、分
        start_datetime_list = []
        end_datetime_list = []

        start_datetime_list.extend(start_date.split('-'))
        start_datetime_list.extend(start_time.split(':'))
        end_datetime_list.extend(end_date.split('-'))
        end_datetime_list.extend((end_time.split(':')))

        # 转换类型
        start_datetime_list = sdt = list(map(int, start_datetime_list))
        end_datetime_list = edt = list(map(int, end_datetime_list))

        # 计算时间差
        start_datetime = datetime(sdt[0], sdt[1], sdt[2], sdt[3], sdt[4])
        end_datetime = datetime(edt[0], edt[1], edt[2], edt[3], edt[4])
        result = str(end_datetime - start_datetime)

        # result E.g -> 4:49:00
        list_ = result.split(':')
        text = f'{list_[0]}h{list_[1]}min'

        return text

    @timing
    def main(self):
        """程序入口"""
        # 检查充放电数据文件是否正确存放
        self.inspection_data_file()

        # 收集充放电数据
        self.collect_charge_data()
        self.collect_discharge_data()

        # 数据量不达标没必要进行存储工作
        if len(self.charge_datas) != self.data_item_number * self.cycles \
                or len(self.discharge_datas) != self.data_item_number * self.cycles:
            print("充放电数据收集异常，请注意输出信息来查询错误，程序不执行存储操作.\n")
            return

        # 保存
        self.save_to_excel()


if __name__ == '__main__':
    item = Discharge()
    item.main()